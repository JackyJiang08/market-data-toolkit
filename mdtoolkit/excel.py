"""Output layer: formatted Excel workbooks and the tidy long table.

Three deliverables, each for a different consumer:
  - per-company workbook  -> a human analyst drilling into one name
  - master workbook       -> quick cross-company comparison
  - tidy long table       -> databases / BI / pivot analysis
"""

from __future__ import annotations

import logging
import os
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config, lineage
from .company import CompanyData

LOG = logging.getLogger("mdtoolkit.excel")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


# --------------------------------------------------------------------------- #
# Formatting helpers
# --------------------------------------------------------------------------- #
def _format_sheet(worksheet, *, freeze: str = "B2") -> None:
    """Header styling, frozen panes, and content-based column widths."""
    worksheet.freeze_panes = freeze
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_cells in worksheet.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None),
                      default=10)
        worksheet.column_dimensions[letter].width = min(max(longest + 2, 12), 48)


def fmt_market_cap(value: Optional[float]) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "N/A"
    for unit, size in (("T", 1e12), ("B", 1e9), ("M", 1e6), ("K", 1e3)):
        if abs(value) >= size:
            return f"{value / size:.2f}{unit}"
    return f"{value:.0f}"


def _pct(x: Optional[float]) -> str:
    return f"{x*100:.3f}%" if x is not None and pd.notna(x) else "N/A"


# --------------------------------------------------------------------------- #
# Per-company workbook
# --------------------------------------------------------------------------- #
def write_company_workbook(data: CompanyData, years: int) -> str:
    path = os.path.join(config.OUTPUT_DIR, f"{data.ticker}_data.xlsx")

    summary = pd.DataFrame({
        "Field": [
            "Ticker", "Company", "Sector", "Industry", "Currency",
            "Market Cap", "Market Cap (raw)",
            "Shares Outstanding (traded class)",
            "Shares Outstanding (reference, mktcap/price)",
            "Last Closing Price", "Dividend Rate (annual)", "Dividend Yield",
            "Financials Window",
            "Extracted (as-of)", "Equity Source", "Rates Source",
        ],
        "Value": [
            data.ticker, data.name, data.sector, data.industry, data.currency,
            fmt_market_cap(data.market_cap), data.market_cap,
            data.shares_traded_class, data.reference_shares,
            data.last_close, data.dividend_rate,
            f"{data.dividend_yield:.2f}%" if data.dividend_yield else "N/A",
            f"{years}y",
            data.as_of, lineage.EQUITY_SOURCE, lineage.RATES_SOURCE,
        ],
    })

    credit_df = pd.DataFrame()
    if data.credit is not None:
        c = data.credit
        credit_df = pd.DataFrame({
            "Field": [
                "Model", "Equity E (latest)", "Default-point Debt D",
                "Risk-free r (1Y Treasury)", "Horizon T (years)",
                "Equity Volatility (annual)",
                "Asset Value V0", "Asset Volatility (annual)",
                "Distance to Default", "Default Probability (PD)", "Note",
            ],
            "Value": [
                c.model,
                data.panel["MarketCap_E"].dropna().iloc[-1] if not data.panel.empty else None,
                data.panel["DefaultPointDebt_D"].dropna().iloc[-1]
                if not data.panel.empty and data.panel["DefaultPointDebt_D"].notna().any() else None,
                _pct(data.panel["RiskFree_R"].dropna().iloc[-1])
                if not data.panel.empty and data.panel["RiskFree_R"].notna().any() else "N/A",
                config.HORIZON_YEARS,
                _pct(_equity_vol(data)),
                c.asset_value, _pct(c.asset_vol),
                round(c.distance_to_default, 4) if c.distance_to_default is not None else None,
                _pct(c.default_probability), c.note,
            ],
        })

    sheets: list[tuple[str, pd.DataFrame, bool]] = [
        ("Summary", summary, False),
        ("Credit Inputs & Estimate", credit_df, False),
        ("Aligned Panel", data.panel, True),
        ("Debt & Liabilities", data.debt_schedule, True),
        ("Price History", data.prices, True),
        ("Dividends", data.dividends, True),
        ("Q Income Statement", data.q_income, True),
        ("Q Balance Sheet", data.q_balance, True),
        ("Q Cash Flow", data.q_cashflow, True),
        ("Annual Income Statement", data.a_income, True),
        ("Annual Balance Sheet", data.a_balance, True),
        ("Annual Cash Flow", data.a_cashflow, True),
    ]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df, with_index in sheets:
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=name[:31], index=with_index)
        for name, df, _ in sheets:
            if df is None or df.empty:
                continue
            _format_sheet(writer.book[name[:31]])
    LOG.info("  -> saved %s", os.path.basename(path))
    return path


def _equity_vol(data: CompanyData) -> Optional[float]:
    if data.panel.empty or "EquityLogReturn" not in data.panel:
        return None
    import numpy as np
    rets = data.panel["EquityLogReturn"].dropna()
    if len(rets) < 6:
        return None
    return float(rets.std() * np.sqrt(config.TRADING_DAYS_PER_YEAR))


# --------------------------------------------------------------------------- #
# Master workbook
# --------------------------------------------------------------------------- #
def write_master_workbook(companies: list[CompanyData], rates: pd.DataFrame) -> str:
    path = os.path.join(config.OUTPUT_DIR, "_MASTER_summary.xlsx")

    summary = pd.DataFrame([{
        "Ticker": c.ticker,
        "Company": c.name,
        "Sector": c.sector,
        "Market Cap": fmt_market_cap(c.market_cap),
        "Market Cap (raw)": c.market_cap,
        "Reference Shares": c.reference_shares,
        "Last Close": c.last_close,
        "Dividend Rate": c.dividend_rate,
        "Dividend Yield (%)": c.dividend_yield,
    } for c in companies])

    debt_rows = []
    for c in companies:
        if c.debt_schedule.empty:
            continue
        latest = c.debt_schedule.columns[0]
        row = {"Ticker": c.ticker, "Period": latest}
        row.update(c.debt_schedule[latest].to_dict())
        debt_rows.append(row)
    debt_latest = pd.DataFrame(debt_rows)

    credit_rows = []
    for c in companies:
        if c.credit is None:
            continue
        E = c.panel["MarketCap_E"].dropna().iloc[-1] if not c.panel.empty else None
        D = (c.panel["DefaultPointDebt_D"].dropna().iloc[-1]
             if not c.panel.empty and c.panel["DefaultPointDebt_D"].notna().any() else None)
        r = (c.panel["RiskFree_R"].dropna().iloc[-1]
             if not c.panel.empty and c.panel["RiskFree_R"].notna().any() else None)
        credit_rows.append({
            "Ticker": c.ticker,
            "Model": c.credit.model,
            "Equity_E": E,
            "DefaultPointDebt_D": D,
            "RiskFree_r": r,
            "EquityVol": _equity_vol(c),
            "AssetValue_V0": c.credit.asset_value,
            "AssetVol_sigmaV": c.credit.asset_vol,
            "DistanceToDefault": c.credit.distance_to_default,
            "DefaultProbability_PD": c.credit.default_probability,
        })
    credit_df = pd.DataFrame(credit_rows)

    rates_display = pd.DataFrame()
    if rates is not None and not rates.empty:
        rates_display = rates.rename(columns=config.FRED_SERIES)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Company Summary", index=False)
        _format_sheet(writer.book["Company Summary"])
        if not credit_df.empty:
            credit_df.to_excel(writer, sheet_name="Credit Summary", index=False)
            _format_sheet(writer.book["Credit Summary"])
        if not debt_latest.empty:
            debt_latest.to_excel(writer, sheet_name="Debt & Liab (latest)", index=False)
            _format_sheet(writer.book["Debt & Liab (latest)"])
        if not rates_display.empty:
            rates_display.to_excel(writer, sheet_name="Macro Rates", index=False)
            _format_sheet(writer.book["Macro Rates"])
    LOG.info("master workbook -> %s", os.path.basename(path))
    return path
