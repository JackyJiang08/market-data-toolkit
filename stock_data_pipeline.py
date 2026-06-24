#!/usr/bin/env python3
"""
Equity & Rates Data Pipeline
============================================================================

Downloads, for a configurable universe of tickers, a curated set of equity
and macro-rate data and writes one formatted Excel workbook per company plus
a consolidated master workbook.

Per company (equity data, source: Yahoo Finance via `yfinance`)
    - Closing price history (daily OHLC + Volume)
    - Dividend-adjusted close + dividend adjustment factor
    - Dividend cash events
    - Market cap (current + estimated history) and shares outstanding
    - Quarterly & annual financial statements (income / balance / cash flow),
      trimmed to a trailing N-year window
    - Extracted debt & liability schedule:
        Total Debt, Total Liabilities,
        Short-term / current Debt, Short-term / current Liabilities,
        Long-term Debt, Long-term Liabilities

Macro rates (source: FRED, St. Louis Fed -- same data as Federal Reserve H.15)
    - DGS1 : 1-Year Treasury Constant Maturity Rate
    - SOFR : Secured Overnight Financing Rate

Design notes
    - Network access to Yahoo is rate-limited; every Yahoo call goes through a
      retry-with-exponential-backoff wrapper and a polite inter-request delay.
    - Balance-sheet line items are resolved against a list of candidate labels
      so the pipeline survives yfinance schema drift.
    - All failures are isolated per ticker; one bad symbol never aborts the run.

Usage
    python3 stock_data_pipeline.py                 # default universe, 2y
    python3 stock_data_pipeline.py AAPL MSFT       # custom tickers
    python3 stock_data_pipeline.py --years 3       # custom financials window
    python3 stock_data_pipeline.py --no-rates      # skip macro-rate download
"""

from __future__ import annotations

import argparse
import logging
import os
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from functools import wraps
from io import StringIO
from typing import Callable, Iterable, Optional, Sequence

try:
    import pandas as pd
    import requests
    import yfinance as yf
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        f"Missing dependency ({exc.name}). Install with:\n"
        "    pip3 install --upgrade yfinance openpyxl pandas requests"
    )


# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

DEFAULT_TICKERS: tuple[str, ...] = (
    "COST", "KO", "DELL", "ORCL", "PNC",
    "WMT", "INTU", "AMZN", "T", "KHC",
)

# FRED series id -> human label. These are the H.15 / SOFR releases.
FRED_SERIES: dict[str, str] = {
    "DGS1": "1-Year Treasury Constant Maturity Rate (%)",
    "SOFR": "Secured Overnight Financing Rate (%)",
}

# Candidate yfinance balance-sheet row labels for each metric we want, in
# priority order. The first label that exists in the frame wins. Matching is
# case- and whitespace-insensitive.
BALANCE_SHEET_MAP: dict[str, tuple[str, ...]] = {
    "Total Debt": ("Total Debt",),
    "Total Liabilities": (
        "Total Liabilities Net Minority Interest",
        "Total Liab",
        "Total Liabilities",
    ),
    "Short-term / Current Debt": (
        "Current Debt And Capital Lease Obligation",
        "Current Debt",
        "Short Term Debt",
        "Short Long Term Debt",
    ),
    "Short-term / Current Liabilities": (
        "Current Liabilities",
        "Total Current Liabilities",
    ),
    "Long-term Debt": (
        "Long Term Debt And Capital Lease Obligation",
        "Long Term Debt",
    ),
    "Long-term Liabilities": (
        "Total Non Current Liabilities Net Minority Interest",
        "Total Non Current Liabilities",
        "Non Current Liabilities",
    ),
}

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

# Network politeness / resilience
MAX_RETRIES = 4
BACKOFF_BASE_SECONDS = 2.0
INTER_TICKER_DELAY_SECONDS = 1.5
REQUEST_TIMEOUT = 20

# --- Data provenance / lineage --------------------------------------------- #
# Captured once at import so every record in a run shares one extraction stamp.
RUN_TIMESTAMP = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
EQUITY_SOURCE = f"Yahoo Finance (yfinance {getattr(yf, '__version__', '?')})"
RATES_SOURCE = "FRED / Federal Reserve H.15 (DGS1, SOFR)"

LOG = logging.getLogger("pipeline")


@dataclass
class RunConfig:
    tickers: Sequence[str]
    years: int = 2
    include_rates: bool = True
    output_dir: str = OUTPUT_DIR

    @property
    def cutoff_date(self) -> datetime:
        return datetime.now() - timedelta(days=365 * self.years + 7)


# --------------------------------------------------------------------------- #
# Resilience helpers
# --------------------------------------------------------------------------- #

def with_retry(label: str) -> Callable:
    """Retry a Yahoo call with exponential backoff on any exception."""
    def decorator(fn: Callable) -> Callable:
        @wraps(fn)
        def wrapper(*args, **kwargs):
            last_exc: Optional[Exception] = None
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    return fn(*args, **kwargs)
                except Exception as exc:  # noqa: BLE001 - intentional broad catch
                    last_exc = exc
                    wait = BACKOFF_BASE_SECONDS * (2 ** (attempt - 1))
                    LOG.warning(
                        "%s failed (attempt %d/%d): %s -- retrying in %.0fs",
                        label, attempt, MAX_RETRIES, exc, wait,
                    )
                    time.sleep(wait)
            LOG.error("%s permanently failed: %s", label, last_exc)
            raise last_exc  # type: ignore[misc]
        return wrapper
    return decorator


def _norm(s: str) -> str:
    return "".join(str(s).lower().split())


def pick_row(frame: pd.DataFrame, candidates: Iterable[str]) -> Optional[pd.Series]:
    """Return the first frame row whose index matches a candidate label."""
    if frame is None or frame.empty:
        return None
    norm_index = {_norm(idx): idx for idx in frame.index}
    for cand in candidates:
        hit = norm_index.get(_norm(cand))
        if hit is not None:
            return frame.loc[hit]
    return None


def trim_to_window(frame: pd.DataFrame, cutoff: datetime) -> pd.DataFrame:
    """Keep statement columns (period-end dates) within the trailing window.

    yfinance statements have period-end dates as columns. Keep those >= cutoff
    but always retain at least the two most recent periods.
    """
    if frame is None or frame.empty:
        return pd.DataFrame()
    cols = list(frame.columns)
    try:
        keep = [c for c in cols if pd.to_datetime(c) >= pd.Timestamp(cutoff)]
    except Exception:  # noqa: BLE001
        keep = cols
    if len(keep) < 2:
        keep = cols[:2]
    return frame[keep]


# --------------------------------------------------------------------------- #
# Macro rates (FRED)
# --------------------------------------------------------------------------- #

def fetch_fred_series(series_id: str, start: datetime, end: datetime) -> pd.DataFrame:
    """Download one FRED series as a tidy [Date, value] frame."""
    url = (
        "https://fred.stlouisfed.org/graph/fredgraph.csv"
        f"?id={series_id}&cosd={start:%Y-%m-%d}&coed={end:%Y-%m-%d}"
    )
    resp = requests.get(url, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    df = pd.read_csv(StringIO(resp.text))
    # FRED columns: observation_date (or DATE), <SERIES_ID>
    date_col = df.columns[0]
    df = df.rename(columns={date_col: "Date"})
    df["Date"] = pd.to_datetime(df["Date"])
    df[series_id] = pd.to_numeric(df[series_id], errors="coerce")
    return df.dropna(subset=[series_id]).reset_index(drop=True)


def build_rates_frame(cfg: RunConfig) -> pd.DataFrame:
    """Combined daily frame of all configured FRED rate series."""
    end = datetime.now()
    start = end - timedelta(days=365 * cfg.years + 7)
    merged: Optional[pd.DataFrame] = None
    for series_id, label in FRED_SERIES.items():
        try:
            df = fetch_fred_series(series_id, start, end)[["Date", series_id]]
            df = df.rename(columns={series_id: label})
            merged = df if merged is None else merged.merge(df, on="Date", how="outer")
            LOG.info("FRED %s: %d observations", series_id, len(df))
        except Exception as exc:  # noqa: BLE001
            LOG.error("FRED %s failed: %s", series_id, exc)
    if merged is None:
        return pd.DataFrame()
    return merged.sort_values("Date").reset_index(drop=True)


# --------------------------------------------------------------------------- #
# Equity data (Yahoo / yfinance)
# --------------------------------------------------------------------------- #

@dataclass
class CompanyData:
    ticker: str
    as_of: str = RUN_TIMESTAMP
    name: str = ""
    currency: str = ""
    sector: str = ""
    industry: str = ""
    market_cap: Optional[float] = None
    shares_outstanding: Optional[float] = None
    implied_shares: Optional[float] = None
    last_close: Optional[float] = None
    dividend_rate: Optional[float] = None
    dividend_yield: Optional[float] = None
    prices: pd.DataFrame = field(default_factory=pd.DataFrame)
    cap_history: pd.DataFrame = field(default_factory=pd.DataFrame)
    dividends: pd.DataFrame = field(default_factory=pd.DataFrame)
    debt_schedule: pd.DataFrame = field(default_factory=pd.DataFrame)
    q_income: pd.DataFrame = field(default_factory=pd.DataFrame)
    q_balance: pd.DataFrame = field(default_factory=pd.DataFrame)
    q_cashflow: pd.DataFrame = field(default_factory=pd.DataFrame)
    a_income: pd.DataFrame = field(default_factory=pd.DataFrame)
    a_balance: pd.DataFrame = field(default_factory=pd.DataFrame)
    a_cashflow: pd.DataFrame = field(default_factory=pd.DataFrame)


@with_retry("Ticker.info")
def _get_info(tk: "yf.Ticker") -> dict:
    return tk.info or {}


@with_retry("Ticker.history")
def _get_history(tk: "yf.Ticker", start: datetime) -> pd.DataFrame:
    return tk.history(start=start.strftime("%Y-%m-%d"),
                      auto_adjust=False, actions=True)


@with_retry("Ticker.statements")
def _get_statements(tk: "yf.Ticker") -> dict[str, pd.DataFrame]:
    return {
        "q_income": tk.quarterly_income_stmt,
        "q_balance": tk.quarterly_balance_sheet,
        "q_cashflow": tk.quarterly_cashflow,
        "a_income": tk.income_stmt,
        "a_balance": tk.balance_sheet,
        "a_cashflow": tk.cashflow,
    }


def build_debt_schedule(balance: pd.DataFrame) -> pd.DataFrame:
    """Extract the requested debt & liability line items as a tidy time series."""
    if balance is None or balance.empty:
        return pd.DataFrame()
    rows: dict[str, pd.Series] = {}
    for metric, candidates in BALANCE_SHEET_MAP.items():
        series = pick_row(balance, candidates)
        if series is not None:
            rows[metric] = series
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows).T  # metrics as rows, period-end dates as columns
    out.columns = [pd.to_datetime(c).strftime("%Y-%m-%d") for c in out.columns]
    return out


def fetch_company(ticker: str, cfg: RunConfig) -> Optional[CompanyData]:
    ticker = ticker.strip().upper()
    if not ticker:
        return None
    LOG.info("=== %s ===", ticker)
    tk = yf.Ticker(ticker)
    data = CompanyData(ticker=ticker)

    # --- company info ---
    try:
        info = _get_info(tk)
    except Exception:  # noqa: BLE001
        info = {}
    data.name = info.get("longName") or info.get("shortName") or ticker
    data.currency = info.get("currency", "")
    data.sector = info.get("sector", "")
    data.industry = info.get("industry", "")
    data.market_cap = info.get("marketCap")
    data.shares_outstanding = info.get("sharesOutstanding")
    # Total implied shares (covers all classes incl. non-traded), keeps
    # market cap = shares * price consistent for dual-class names (e.g. DELL).
    data.implied_shares = info.get("impliedSharesOutstanding")
    data.dividend_rate = info.get("dividendRate")
    data.dividend_yield = info.get("dividendYield")

    # --- price history (within window) ---
    try:
        prices = _get_history(tk, cfg.cutoff_date)
    except Exception:  # noqa: BLE001
        prices = pd.DataFrame()
    if not prices.empty:
        prices.index = prices.index.tz_localize(None)
        # Dividend adjustment factor: AdjClose / Close (cumulative div+split factor)
        if "Adj Close" in prices and "Close" in prices:
            prices["Div/Split Adj Factor"] = (
                prices["Adj Close"] / prices["Close"]
            ).round(6)
        data.prices = prices
        data.last_close = float(prices["Close"].iloc[-1])
        LOG.info("  prices: %d rows (%s -> %s)", len(prices),
                 prices.index.min().date(), prices.index.max().date())

    # --- dividends ---
    if not prices.empty and "Dividends" in prices:
        divs = prices.loc[prices["Dividends"] > 0, ["Dividends"]].copy()
        if not divs.empty:
            divs.index.name = "Date"
            data.dividends = divs

    # --- estimated market-cap history (prefer implied total shares) ---
    cap_shares = data.implied_shares or data.shares_outstanding
    if not prices.empty and cap_shares:
        data.cap_history = pd.DataFrame({
            "Close": prices["Close"],
            "SharesUsed": cap_shares,
            "EstMarketCap": prices["Close"] * cap_shares,
        })

    # --- financial statements ---
    try:
        stmts = _get_statements(tk)
    except Exception:  # noqa: BLE001
        stmts = {}
    data.q_income = trim_to_window(stmts.get("q_income"), cfg.cutoff_date)
    data.q_balance = trim_to_window(stmts.get("q_balance"), cfg.cutoff_date)
    data.q_cashflow = trim_to_window(stmts.get("q_cashflow"), cfg.cutoff_date)
    data.a_income = trim_to_window(stmts.get("a_income"), cfg.cutoff_date)
    data.a_balance = trim_to_window(stmts.get("a_balance"), cfg.cutoff_date)
    data.a_cashflow = trim_to_window(stmts.get("a_cashflow"), cfg.cutoff_date)

    # Prefer quarterly balance sheet for the debt schedule; fall back to annual.
    data.debt_schedule = build_debt_schedule(
        data.q_balance if not data.q_balance.empty else data.a_balance
    )
    if not data.debt_schedule.empty:
        LOG.info("  debt schedule: %d metrics x %d periods",
                 *data.debt_schedule.shape)
    return data


# --------------------------------------------------------------------------- #
# Excel output
# --------------------------------------------------------------------------- #

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INDEX_FONT = Font(bold=True)


def _format_sheet(worksheet, *, freeze: str = "B2", index_width: int = 38) -> None:
    """Apply header styling, freeze panes, and content-based column widths.

    Every column (including the first) is sized to its widest value, clamped to
    a sensible range, so dates and long labels are never truncated to '####'.
    """
    worksheet.freeze_panes = freeze
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_cells in worksheet.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None),
                      default=10)
        worksheet.column_dimensions[letter].width = min(max(longest + 2, 12), 48)


def write_company_workbook(data: CompanyData, cfg: RunConfig) -> str:
    path = os.path.join(cfg.output_dir, f"{data.ticker}_data.xlsx")
    fmt = fmt_market_cap

    summary = pd.DataFrame({
        "Field": [
            "Ticker", "Company", "Sector", "Industry", "Currency",
            "Market Cap", "Market Cap (raw)",
            "Shares Outstanding (traded class)",
            "Shares Outstanding (implied total)",
            "Last Closing Price", "Dividend Rate (annual)", "Dividend Yield",
            "Financials Window",
            "Extracted (as-of)", "Equity Source", "Rates Source",
        ],
        "Value": [
            data.ticker, data.name, data.sector, data.industry, data.currency,
            fmt(data.market_cap), data.market_cap,
            data.shares_outstanding, data.implied_shares,
            data.last_close, data.dividend_rate,
            f"{data.dividend_yield:.2f}%" if data.dividend_yield else "N/A",
            f"{cfg.years}y",
            data.as_of, EQUITY_SOURCE,
            RATES_SOURCE if cfg.include_rates else "(skipped)",
        ],
    })

    sheets: list[tuple[str, pd.DataFrame, bool]] = [
        ("Summary", summary, False),
        ("Debt & Liabilities", data.debt_schedule, True),
        ("Price History", data.prices, True),
        ("MarketCap History", data.cap_history, True),
        ("Dividends", data.dividends, True),
        ("Q Income Statement", data.q_income, True),
        ("Q Balance Sheet", data.q_balance, True),
        ("Q Cash Flow", data.q_cashflow, True),
        ("Annual Income Statement", data.a_income, True),
        ("Annual Balance Sheet", data.a_balance, True),
        ("Annual Cash Flow", data.a_cashflow, True),
    ]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df, with_index in sheets:
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=name[:31], index=with_index)
        for name, df, _ in sheets:
            if df is None or df.empty:
                continue
            _format_sheet(writer.book[name[:31]])
    LOG.info("  -> saved %s", os.path.basename(path))
    return path


def write_master_workbook(companies: list[CompanyData], rates: pd.DataFrame,
                          cfg: RunConfig) -> str:
    path = os.path.join(cfg.output_dir, "_MASTER_summary.xlsx")

    summary = pd.DataFrame([{
        "Ticker": c.ticker,
        "Company": c.name,
        "Sector": c.sector,
        "Currency": c.currency,
        "Market Cap": fmt_market_cap(c.market_cap),
        "Market Cap (raw)": c.market_cap,
        "Shares Outstanding": c.shares_outstanding,
        "Last Close": c.last_close,
        "Dividend Rate": c.dividend_rate,
        "Dividend Yield": (round(c.dividend_yield, 4)
                           if c.dividend_yield else None),
    } for c in companies])

    # Latest debt & liabilities, one row per company.
    debt_rows = []
    for c in companies:
        if c.debt_schedule.empty:
            continue
        latest = c.debt_schedule.columns[0]
        row = {"Ticker": c.ticker, "Period": latest}
        row.update(c.debt_schedule[latest].to_dict())
        debt_rows.append(row)
    debt_latest = pd.DataFrame(debt_rows)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Company Summary", index=False)
        _format_sheet(writer.book["Company Summary"], index_width=14)
        if not debt_latest.empty:
            debt_latest.to_excel(writer, sheet_name="Debt & Liab (latest)",
                                 index=False)
            _format_sheet(writer.book["Debt & Liab (latest)"], index_width=10)
        if not rates.empty:
            rates.to_excel(writer, sheet_name="Macro Rates", index=False)
            _format_sheet(writer.book["Macro Rates"], index_width=12)
    LOG.info("master workbook -> %s", os.path.basename(path))
    return path


def fmt_market_cap(value: Optional[float]) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "N/A"
    for unit, size in (("T", 1e12), ("B", 1e9), ("M", 1e6), ("K", 1e3)):
        if abs(value) >= size:
            return f"{value / size:.2f}{unit}"
    return f"{value:.0f}"


# --------------------------------------------------------------------------- #
# Tidy long table (machine-readable consolidation of every company + rates)
# --------------------------------------------------------------------------- #

LONG_COLUMNS = ["Ticker", "AsOf", "Category", "Period", "Metric", "Value"]


def _melt_statement(df: pd.DataFrame, ticker: str, as_of: str,
                    category: str) -> pd.DataFrame:
    """Melt a statement frame (rows=line items, cols=period dates) to long."""
    if df is None or df.empty:
        return pd.DataFrame(columns=LONG_COLUMNS)
    t = df.copy()
    t.index.name = "Metric"
    long = t.reset_index().melt(id_vars="Metric", var_name="Period",
                                value_name="Value")
    long["Period"] = pd.to_datetime(long["Period"], errors="coerce").dt.strftime("%Y-%m-%d")
    long["Ticker"], long["AsOf"], long["Category"] = ticker, as_of, category
    return long[LONG_COLUMNS]


def _melt_timeseries(df: pd.DataFrame, ticker: str, as_of: str,
                     category: str) -> pd.DataFrame:
    """Melt a time-indexed frame (rows=dates, cols=metrics) to long."""
    if df is None or df.empty:
        return pd.DataFrame(columns=LONG_COLUMNS)
    t = df.copy()
    t.index.name = "Period"
    long = t.reset_index().melt(id_vars="Period", var_name="Metric",
                                value_name="Value")
    long["Period"] = pd.to_datetime(long["Period"], errors="coerce").dt.strftime("%Y-%m-%d")
    long["Ticker"], long["AsOf"], long["Category"] = ticker, as_of, category
    return long[LONG_COLUMNS]


def build_long_table(companies: list[CompanyData], rates: pd.DataFrame) -> pd.DataFrame:
    """One tidy fact table: [Ticker, AsOf, Category, Period, Metric, Value]."""
    parts: list[pd.DataFrame] = []
    for c in companies:
        # Point-in-time company snapshot.
        snapshot = {
            "MarketCap": c.market_cap,
            "SharesOutstanding": c.shares_outstanding,
            "LastClose": c.last_close,
            "DividendRate": c.dividend_rate,
            "DividendYield": c.dividend_yield,
        }
        snap_df = pd.DataFrame(
            [(c.ticker, c.as_of, "company_info", c.as_of[:10], k, v)
             for k, v in snapshot.items() if v is not None],
            columns=LONG_COLUMNS,
        )
        parts.append(snap_df)
        parts.append(_melt_timeseries(c.prices, c.ticker, c.as_of, "price"))
        parts.append(_melt_timeseries(c.dividends, c.ticker, c.as_of, "dividend"))
        parts.append(_melt_statement(c.debt_schedule, c.ticker, c.as_of, "debt_schedule"))
        parts.append(_melt_statement(c.q_income, c.ticker, c.as_of, "income_statement (Q)"))
        parts.append(_melt_statement(c.q_balance, c.ticker, c.as_of, "balance_sheet (Q)"))
        parts.append(_melt_statement(c.q_cashflow, c.ticker, c.as_of, "cash_flow (Q)"))
        parts.append(_melt_statement(c.a_income, c.ticker, c.as_of, "income_statement (A)"))
        parts.append(_melt_statement(c.a_balance, c.ticker, c.as_of, "balance_sheet (A)"))
        parts.append(_melt_statement(c.a_cashflow, c.ticker, c.as_of, "cash_flow (A)"))

    if rates is not None and not rates.empty:
        r = rates.rename(columns={"Date": "Period"}).copy()
        r["Period"] = pd.to_datetime(r["Period"]).dt.strftime("%Y-%m-%d")
        r_long = r.melt(id_vars="Period", var_name="Metric", value_name="Value")
        r_long["Ticker"], r_long["AsOf"], r_long["Category"] = "MACRO", RUN_TIMESTAMP, "rate"
        parts.append(r_long[LONG_COLUMNS])

    if not parts:
        return pd.DataFrame(columns=LONG_COLUMNS)
    out = pd.concat(parts, ignore_index=True)
    out["Value"] = pd.to_numeric(out["Value"], errors="coerce")
    out = out.dropna(subset=["Value"])
    return out.sort_values(["Ticker", "Category", "Period", "Metric"]).reset_index(drop=True)


def write_long_table(long_df: pd.DataFrame, cfg: RunConfig) -> None:
    if long_df.empty:
        LOG.warning("long table is empty -- nothing written")
        return
    csv_path = os.path.join(cfg.output_dir, "all_companies_long.csv")
    long_df.to_csv(csv_path, index=False)
    LOG.info("tidy long table -> %s (%d rows)", os.path.basename(csv_path), len(long_df))
    try:
        pq_path = os.path.join(cfg.output_dir, "all_companies_long.parquet")
        long_df.to_parquet(pq_path, index=False)
        LOG.info("tidy long table -> %s", os.path.basename(pq_path))
    except Exception as exc:  # noqa: BLE001 - parquet engine optional
        LOG.info("parquet skipped (%s); CSV is available", exc)


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #

def run(cfg: RunConfig) -> None:
    os.makedirs(cfg.output_dir, exist_ok=True)
    LOG.info("Universe: %s | window: %dy | rates: %s",
             ", ".join(cfg.tickers), cfg.years, cfg.include_rates)

    rates = build_rates_frame(cfg) if cfg.include_rates else pd.DataFrame()

    companies: list[CompanyData] = []
    for i, ticker in enumerate(cfg.tickers):
        try:
            data = fetch_company(ticker, cfg)
            if data is not None:
                write_company_workbook(data, cfg)
                companies.append(data)
        except Exception as exc:  # noqa: BLE001
            LOG.error("Ticker %s aborted: %s", ticker, exc)
        if i < len(cfg.tickers) - 1:
            time.sleep(INTER_TICKER_DELAY_SECONDS)

    if companies:
        write_master_workbook(companies, rates, cfg)
        write_long_table(build_long_table(companies, rates), cfg)

    LOG.info("Done. %d/%d companies succeeded. Output: %s",
             len(companies), len(cfg.tickers), cfg.output_dir)


def parse_args(argv: Optional[Sequence[str]] = None) -> RunConfig:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("tickers", nargs="*", help="Ticker symbols (default: built-in universe)")
    p.add_argument("--years", type=int, default=2, help="Financials/history window in years (default 2)")
    p.add_argument("--no-rates", action="store_true", help="Skip FRED macro-rate download")
    args = p.parse_args(argv)
    tickers = [t.upper() for t in args.tickers] or list(DEFAULT_TICKERS)
    return RunConfig(tickers=tickers, years=args.years, include_rates=not args.no_rates)


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%H:%M:%S",
    )
    run(parse_args())


if __name__ == "__main__":
    main()
